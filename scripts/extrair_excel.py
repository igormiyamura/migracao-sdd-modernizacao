#!/usr/bin/env python
"""Pré-processamento determinístico de uma planilha Excel (.xlsx/.xlsm) para
extração de discovery -- ver
.claude/skills/migrar-processo/referencias/extracao-monolitos.md, seção 1.

Lê o arquivo com openpyxl (nunca abre o Excel) e produz um resumo YAML
compacto: fórmulas por aba, ranges nomeados, validação de dados e
formatação condicional. Linhas de dado puro (sem fórmula) em blocos grandes
são resumidas (schema + amostra), não despejadas por inteiro -- é
exatamente o tipo de conteúdo que não cabe carregar inteiro no contexto do
agente e raramente carrega lógica de negócio linha a linha.

Uso: python extrair_excel.py <caminho_planilha.xlsx> [--limiar-resumo N]
Saída: YAML no stdout.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.utils import get_column_letter

LIMIAR_RESUMO_PADRAO = 20


def _celulas_nao_vazias(ws):
    for linha in ws.iter_rows():
        for celula in linha:
            if celula.value is not None:
                yield celula


def _extrair_aba(ws, limiar_resumo: int) -> dict:
    formulas = []
    valores_literais = []

    for celula in _celulas_nao_vazias(ws):
        endereco = f"{get_column_letter(celula.column)}{celula.row}"
        if isinstance(celula.value, str) and celula.value.startswith("="):
            formulas.append({"celula": endereco, "formula": celula.value})
        else:
            valores_literais.append(
                {"celula": endereco, "linha": celula.row, "valor": celula.value}
            )

    bloco_valores: dict
    if len(valores_literais) <= limiar_resumo:
        bloco_valores = {"tipo": "completo", "celulas": valores_literais}
    else:
        linhas = sorted({v["linha"] for v in valores_literais})
        amostra = [v for v in valores_literais if v["linha"] in linhas[:3]]
        bloco_valores = {
            "tipo": "resumo",
            "total_celulas_com_valor": len(valores_literais),
            "linhas_com_dado": [linhas[0], linhas[-1]] if linhas else None,
            "amostra_primeiras_linhas": amostra,
            "nota": (
                "bloco de dado grande, resumido -- ler a fonte original "
                "(ou entrada/carteira_*.csv equivalente) se precisar de todos os valores"
            ),
        }

    validacoes = [
        {"range": str(dv.sqref), "tipo": dv.type, "formula": dv.formula1}
        for dv in ws.data_validations.dataValidation
    ]

    formatacoes_condicionais = []
    for cf in ws.conditional_formatting:
        for regra in cf.rules:
            formatacoes_condicionais.append(
                {
                    "range": str(cf.sqref),
                    "tipo": regra.type,
                    "operador": regra.operator,
                    "formula": list(regra.formula) if regra.formula else None,
                }
            )

    return {
        "formulas": formulas,
        "valores": bloco_valores,
        "validacoes_dado": validacoes,
        "formatacao_condicional": formatacoes_condicionais,
    }


def extrair(caminho: Path, limiar_resumo: int) -> dict:
    wb = openpyxl.load_workbook(caminho, data_only=False)

    ranges_nomeados = [
        {"nome": nome, "referencia": destino.value}
        for nome, destino in wb.defined_names.items()
    ]

    abas = {}
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        abas[nome_aba] = {
            "oculta": ws.sheet_state != "visible",
            **_extrair_aba(ws, limiar_resumo),
        }

    return {
        "arquivo": str(caminho),
        "abas": abas,
        "ranges_nomeados": ranges_nomeados,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("caminho", type=Path)
    parser.add_argument("--limiar-resumo", type=int, default=LIMIAR_RESUMO_PADRAO)
    args = parser.parse_args()

    if not args.caminho.exists():
        print(f"erro: arquivo não encontrado: {args.caminho}", file=sys.stderr)
        raise SystemExit(1)

    resultado = extrair(args.caminho, args.limiar_resumo)
    yaml.dump(resultado, sys.stdout, allow_unicode=True, sort_keys=False, width=100)


if __name__ == "__main__":
    main()
